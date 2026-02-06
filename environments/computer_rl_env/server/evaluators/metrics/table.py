import functools
import itertools
import logging
import os
import re
import unicodedata
from typing import Callable, Dict, Iterable, List, Optional, Set, Tuple, Union, cast

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz

from .utils import (
    _match_value_to_rule,
    _read_cell_style,
    load_charts,
    load_filters,
    load_pivot_tables,
    load_rows_or_cols,
    load_sparklines,
    load_xlsx_styles,
    read_cell_value,
)

logger = logging.getLogger("computer_rl_env.server.evaluators.metrics.table")

BOOK = Union[pd.ExcelFile, Workbook, str]
SHEET = Union[pd.DataFrame, Worksheet, List[str]]
RuleValue = Union[str, int, float, bool, List[str], Dict[str, Union[str, int, float, bool]]]
Rule = Dict[str, RuleValue]


def _parse_sheet_idx(
    sheet_idx: Union[int, str],
    result: BOOK,
    expected: Optional[BOOK],
    result_sheet_names: List[Union[str, int]],
    expected_sheet_names: Optional[List[Union[str, int]]],
) -> Tuple[BOOK, str]:
    """Parse sheet index into book and sheet name."""
    if isinstance(sheet_idx, int):
        try:
            if not result_sheet_names or sheet_idx >= len(result_sheet_names):
                logger.error(
                    f"Sheet index {sheet_idx} out of range. Available sheets: {result_sheet_names}"
                )
                index = ""
            else:
                index: str = str(result_sheet_names[sheet_idx])
                logger.debug(f"Sheet index {sheet_idx} resolved to sheet: {index}")
        except Exception as e:
            logger.error(f"Error resolving sheet index {sheet_idx}: {e}")
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RI"):
        try:
            index: str = str(result_sheet_names[int(sheet_idx[2:])])
        except Exception:
            index = ""
        book: BOOK = result
    elif sheet_idx.startswith("RN"):
        index: str = sheet_idx[2:]
        book: BOOK = result
    elif sheet_idx.startswith("EI"):
        try:
            if expected_sheet_names:
                index: str = str(expected_sheet_names[int(sheet_idx[2:])])
            else:
                index = ""
        except Exception:
            index = ""
        book: BOOK = expected if expected else result
    elif sheet_idx.startswith("EN"):
        index: str = sheet_idx[2:]
        book: BOOK = expected if expected else result
    else:
        logger.error("Unrecognized sheet index")
        raise ValueError("Unrecognized sheet index")
    return book, index


def _load_sheet(book: BOOK, index: str) -> Optional[SHEET]:
    """Load a sheet from a workbook."""
    try:
        if isinstance(book, str):
            book_str: str = cast(str, book)
            csv_name: str = "{:}-{:}.csv".format(os.path.splitext(book_str)[0], index)

            try:
                all_lines: List[str] = _safe_read_file(csv_name)
                csv_lines: List[str] = list(
                    itertools.dropwhile(
                        lambda line: len(line) == 0,
                        map(lambda line: line.strip(), reversed(all_lines)),
                    )
                )
                return csv_lines
            except (FileNotFoundError, IOError) as e:
                logger.error(f"Failed to read CSV file {csv_name}: {e}")
                return None
        if isinstance(book, pd.ExcelFile):
            return pd.read_excel(book, index)
        if isinstance(book, Workbook):
            return book[index]
        logger.error("Not supported workbook format")
        raise NotImplementedError("Not supported workbook format")
    except NotImplementedError:
        raise
    except Exception:
        return None


def _safe_read_file(file_path: str) -> List[str]:
    """
    Safely read a file with multiple encoding attempts.

    Args:
        file_path: Path to the file to read

    Returns:
        List of lines from the file

    Raises:
        FileNotFoundError: If file doesn't exist
        IOError: If file cannot be read with any encoding
    """
    # Common encodings to try in order of preference
    encodings = [
        "utf-8",  # Most common modern encoding
        "utf-8-sig",  # UTF-8 with BOM
        "latin-1",  # ISO-8859-1, works with any byte sequence
        "windows-1252",  # Common Windows encoding
        "gbk",  # Chinese encoding
        "cp1251",  # Cyrillic encoding
        "iso-8859-1",  # Alternative latin-1
    ]

    last_error: Optional[Exception] = None

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                lines = f.read().splitlines()
                logger.debug(f"Successfully read file {file_path} with encoding {encoding}")
                return lines
        except UnicodeDecodeError as e:
            last_error = e
            logger.debug(f"Failed to read {file_path} with encoding {encoding}: {e}")
            continue
        except (FileNotFoundError, IOError):
            # These are non-encoding related errors, re-raise immediately
            raise

    # If all encodings fail, try with error handling as last resort
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.read().splitlines()
            logger.warning(f"Read file {file_path} with UTF-8 and error replacement")
            return lines
    except Exception:
        logger.error(f"Failed to read file {file_path} with any encoding. Last error: {last_error}")
        raise IOError(f"Cannot read file {file_path} with any supported encoding") from last_error


def compare_csv(result: Optional[str], expected: Union[str, List[str]], **options: bool) -> float:
    """
    Compare CSV files. If expected is a list, returns 1.0 if result matches any of the expected files.

    Args:
        result: Path to result CSV file
        expected: Path to expected CSV file or list of paths to expected CSV files
        options: Additional options (strict, ignore_case)

    Returns:
        1.0 if result matches expected (or any file in expected list), 0.0 otherwise
    """
    if result is None:
        return 0.0

    try:
        result_lines: List[str] = _safe_read_file(result)
    except (FileNotFoundError, IOError) as e:
        logger.error(f"Failed to read result file {result}: {e}")
        return 0.0

    # Convert expected to list if it's a single string (for backward compatibility)
    if isinstance(expected, str):
        expected_files = [expected]
    else:
        expected_files = expected

    # Try to match against each expected file
    for expected_file in expected_files:
        try:
            expected_lines: List[str] = _safe_read_file(expected_file)

            # Process lines based on options
            current_result_lines: Iterable[str] = result_lines
            current_expected_lines: Iterable[str] = expected_lines

            if not options.get("strict", True):
                current_result_lines = map(str.strip, current_result_lines)
                current_expected_lines = map(str.strip, current_expected_lines)
            if options.get("ignore_case", False):
                current_result_lines = map(str.lower, current_result_lines)
                current_expected_lines = map(str.lower, current_expected_lines)

            # Check if this expected file matches
            if list(current_result_lines) == list(current_expected_lines):
                return 1.0

        except (FileNotFoundError, IOError):
            # If this expected file doesn't exist, continue to next one
            continue

    # No match found
    return 0.0


def compare_table(
    result: Optional[str], expected: Optional[str] = None, **options: RuleValue
) -> float:
    """
    Compare Excel tables with complex rules.

    Args:
        result: Path to result xlsx
        expected: Path to golden xlsx
        options: Dictionary containing 'rules' - list of comparison rules

    Returns:
        The score (1.0 if all rules pass, 0.0 otherwise)
    """

    if result is None:
        logger.error("Result file path is None")
        return 0.0

    # Check if result file exists
    if not os.path.exists(result):
        logger.error(f"Result file not found: {result}")
        return 0.0

    try:
        logger.info(f"Loading result file: {result}")
        xlworkbookr: Workbook = openpyxl.load_workbook(filename=result)
        pdworkbookr = pd.ExcelFile(result)
        logger.info(f"Successfully loaded result file with sheets: {pdworkbookr.sheet_names}")
    except Exception as e:
        logger.error(f"Failed to load result file {result}: {e}")
        return 0.0
    worksheetr_names: List[Union[str, int]] = list(pdworkbookr.sheet_names)

    if expected is not None:
        xlworkbooke: Optional[Workbook] = openpyxl.load_workbook(filename=expected)
        pdworkbooke: Optional[pd.ExcelFile] = pd.ExcelFile(expected)
        worksheete_names: Optional[List[Union[str, int]]] = (
            list(pdworkbooke.sheet_names) if pdworkbooke else None
        )
    else:
        xlworkbooke: Optional[Workbook] = None
        pdworkbooke: Optional[pd.ExcelFile] = None
        worksheete_names: Optional[List[Union[str, int]]] = None

    parse_idx: Callable[[Union[str, int], BOOK, Optional[BOOK]], Tuple[BOOK, str]] = (
        functools.partial(
            _parse_sheet_idx,
            result_sheet_names=worksheetr_names,
            expected_sheet_names=worksheete_names,
        )
    )

    passes = True
    rules_list: List[Rule] = options.get("rules", [])  # type: ignore
    for r in rules_list:
        rule_type = str(r.get("type", ""))
        if rule_type == "sheet_name":
            #  Compare Sheet Names
            metric: bool = worksheetr_names == worksheete_names
            logger.debug(
                "Assertion: %s.sheet_names == %s.sheet_names - %s",
                result,
                expected,
                metric,
            )

        elif rule_type == "sheet_data":
            #  Compare Sheet Data by Internal Value
            error_limit: int = int(r.get("precision", 4))  # type: ignore
            sheet_idx0 = r.get("sheet_idx0", 0)
            sheet_idx1 = r.get("sheet_idx1", 0)
            sheet1_raw = _load_sheet(
                *parse_idx(sheet_idx0, pdworkbookr, pdworkbooke)  # type: ignore
            )
            if sheet1_raw is None or not isinstance(sheet1_raw, pd.DataFrame):
                return 0.0
            sheet2_raw = _load_sheet(
                *parse_idx(sheet_idx1, pdworkbookr, pdworkbooke)  # type: ignore
            )
            if sheet2_raw is None or not isinstance(sheet2_raw, pd.DataFrame):
                return 0.0

            sheet1_df = cast(pd.DataFrame, sheet1_raw).round(error_limit)
            sheet2_df = cast(pd.DataFrame, sheet2_raw).round(error_limit)
            metric: bool = sheet1_df.equals(sheet2_df)
            logger.debug("Sheet1: \n%s", str(sheet1_df))
            logger.debug("Sheet2: \n%s", str(sheet2_df))
            try:
                logger.debug("Sheet1 =v= Sheet2: \n%s", str(sheet1_df == sheet2_df))
            except Exception:
                logger.debug("Sheet1 =/v= Sheet2")
            logger.debug("Assertion: %s =v= %s - %s", sheet_idx0, sheet_idx1, metric)

        elif rule_type == "sheet_print":
            #  Compare Sheet Data by Printed Value
            sheet_idx0 = r.get("sheet_idx0", 0)
            sheet_idx1 = r.get("sheet_idx1", 0)
            sheet1_raw = _load_sheet(
                *parse_idx(sheet_idx0, result, expected)  # type: ignore
            )
            if sheet1_raw is None or not isinstance(sheet1_raw, list):
                return 0.0
            sheet2_raw = _load_sheet(
                *parse_idx(sheet_idx1, result, expected)  # type: ignore
            )
            if sheet2_raw is None or not isinstance(sheet2_raw, list):
                return 0.0

            sheet1_list = cast(List[str], sheet1_raw)
            sheet2_list = cast(List[str], sheet2_raw)
            if r.get("ignore_case", False):
                sheet1_list = [line.lower() for line in sheet1_list]
                sheet2_list = [line.lower() for line in sheet2_list]
            metric: bool = sheet1_list == sheet2_list
            logger.debug("Assertion: %s =p= %s - %s", sheet_idx0, sheet_idx1, metric)

        elif rule_type == "sheet_fuzzy":
            #  Fuzzy Match for Ranges
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            sheet1: Tuple[BOOK, str] = parse_idx(sheet_idx0, result, expected)  # type: ignore
            sheet2: Tuple[BOOK, str] = parse_idx(sheet_idx1, result, expected)  # type: ignore
            total_metric = True
            sub_rules = r.get("rules", [])
            if not isinstance(sub_rules, list):
                sub_rules = []
            for rl_item in sub_rules:
                rl = cast(Rule, rl_item)
                range_val = rl.get("range", [])
                if not isinstance(range_val, (list, str)):
                    continue
                for rng in MultiCellRange(range_val):
                    for cdn in rng.cells:
                        coordinate: str = "{:}{:d}".format(get_column_letter(cdn[1]), cdn[0])
                        value1: str = str(
                            read_cell_value(cast(str, sheet1[0]), sheet1[1], coordinate)
                        )
                        value2: str = str(
                            read_cell_value(cast(str, sheet2[0]), sheet2[1], coordinate)
                        )
                        logger.debug("%s: %s vs %s", cdn, value1, value2)

                        normalization_val = rl.get("normalization", [])
                        normalization = (
                            list(normalization_val) if isinstance(normalization_val, list) else []
                        )
                        for rplc_item in normalization:
                            if isinstance(rplc_item, list) and len(rplc_item) >= 2:
                                value1 = value1.replace(str(rplc_item[0]), str(rplc_item[1]))
                                value2 = value2.replace(str(rplc_item[0]), str(rplc_item[1]))
                        if "trim_leadings" in rl:
                            trim_lead = str(rl.get("trim_leadings", ""))
                            value1 = value1.lstrip(trim_lead)
                            value2 = value2.lstrip(trim_lead)
                        if "trim_trailings" in rl:
                            trim_trail = str(rl.get("trim_trailings", ""))
                            value1 = value1.rstrip(trim_trail)
                            value2 = value2.rstrip(trim_trail)
                        if "ignore_chars" in rl:
                            ignore_chars_val = rl.get("ignore_chars", "")
                            ignore_chars_str = str(ignore_chars_val) if ignore_chars_val else ""
                            ignore_chars: Set[str] = set(ignore_chars_str)
                            value1 = "".join(ch for ch in value1 if ch not in ignore_chars)
                            value2 = "".join(ch for ch in value2 if ch not in ignore_chars)
                        if rl.get("ignore_case", False):
                            value1 = value1.lower()
                            value2 = value2.lower()

                        rl_type = str(rl.get("type", ""))
                        metric_inner: bool = False
                        if rl_type == "includes":
                            metric_inner = value2 in value1
                        elif rl_type == "included_by":
                            metric_inner = value1 in value2
                        elif rl_type == "fuzzy_match":
                            threshold_val = rl.get("threshold", 85.0)
                            threshold = (
                                float(threshold_val)
                                if isinstance(threshold_val, (int, float))
                                else 85.0
                            )
                            metric_inner = fuzz.ratio(value1, value2) >= threshold
                        elif rl_type == "exact_match":
                            metric_inner = value1 == value2
                        total_metric = total_metric and metric_inner

            metric: bool = total_metric
            logger.debug("Assertion: %s =~= %s - %s", sheet_idx0, sheet_idx1, metric)

        elif rule_type == "sparkline":
            #  Compare Sparklines
            sheet_idx0 = r.get("sheet_idx0", 0)
            sheet_idx1 = r.get("sheet_idx1", 0)
            sparkline1: Dict[str, str] = load_sparklines(
                *parse_idx(sheet_idx0, result, expected)  # type: ignore
            )
            sparkline2: Dict[str, str] = load_sparklines(
                *parse_idx(sheet_idx1, result, expected)  # type: ignore
            )
            metric: bool = sparkline1 == sparkline2
            logger.debug(
                "Assertion: %s.sp == %.sp - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "chart":
            #  Compare Charts
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            charts1: Dict[str, Union[str, int, float, List[str]]] = load_charts(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            charts2: Dict[str, Union[str, int, float, List[str]]] = load_charts(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            metric: bool = charts1 == charts2
            logger.debug(
                "Assertion: %s[chart] == %s[chart] - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "style":
            #  Compare Style (Also Conditional Formatiing)
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            sheet_idx1_tuple: Tuple[BOOK, str] = parse_idx(
                sheet_idx0,
                xlworkbookr,
                xlworkbooke,  # type: ignore
            )
            book_name1: str = str(parse_idx(sheet_idx0, result, expected)[0])  # type: ignore
            styles1: Dict[str, List[Union[str, float, bool, None]]] = load_xlsx_styles(
                cast(Workbook, sheet_idx1_tuple[0]),
                sheet_idx1_tuple[1],
                book_name1,
                **r,  # type: ignore
            )

            sheet_idx2_tuple: Tuple[BOOK, str] = parse_idx(
                sheet_idx1,
                xlworkbookr,
                xlworkbooke,  # type: ignore
            )
            book_name2: str = str(parse_idx(sheet_idx1, result, expected)[0])  # type: ignore
            styles2: Dict[str, List[Union[str, float, bool, None]]] = load_xlsx_styles(
                cast(Workbook, sheet_idx2_tuple[0]),
                sheet_idx2_tuple[1],
                book_name2,
                **r,  # type: ignore
            )
            metric: bool = styles1 == styles2
            logger.debug(
                "Assertion: %s.style == %s.style - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "freeze":
            #  Compare Freezing
            sheet_idx0 = r.get("sheet_idx0", 0)
            sheet_idx1 = r.get("sheet_idx1", 0)
            sheet1_raw = _load_sheet(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke)  # type: ignore
            )
            if sheet1_raw is None or not isinstance(sheet1_raw, Worksheet):
                return 0.0
            sheet2_raw = _load_sheet(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke)  # type: ignore
            )
            if sheet2_raw is None or not isinstance(sheet2_raw, Worksheet):
                return 0.0

            sheet1_ws = cast(Worksheet, sheet1_raw)
            sheet2_ws = cast(Worksheet, sheet2_raw)
            metric: bool = sheet1_ws.freeze_panes == sheet2_ws.freeze_panes
            logger.debug(
                "Assertion: %s.freeze(%s) == %s.freeze(%s) - %s",
                sheet_idx0,
                sheet1_ws.freeze_panes,
                sheet_idx1,
                sheet2_ws.freeze_panes,
                metric,
            )

        elif rule_type == "zoom":
            #  Check Zooming
            sheet_idx = r.get("sheet_idx", 0)
            sheet_raw = _load_sheet(
                *parse_idx(sheet_idx, xlworkbookr, xlworkbooke)  # type: ignore
            )
            if sheet_raw is None or not isinstance(sheet_raw, Worksheet):
                return 0.0
            sheet_ws = cast(Worksheet, sheet_raw)
            zoom_scale: float = float(sheet_ws.sheet_view.zoomScale or 100.0)
            metric: bool = _match_value_to_rule(zoom_scale, r)
            logger.debug(
                "Assertion: %s.zoom(%.1f) %s %.1f - %s",
                sheet_idx,
                zoom_scale,
                r.get("method", ""),
                r.get("ref", 0),
                metric,
            )

        elif rule_type == "data_validation":
            #  Check Data Validation
            sheet_idx = cast(Union[int, str], r.get("sheet_idx", 0))
            sheet_raw = _load_sheet(
                *parse_idx(sheet_idx, xlworkbookr, xlworkbooke)  # type: ignore
            )
            if sheet_raw is None or not isinstance(sheet_raw, Worksheet):
                return 0.0
            sheet_ws = cast(Worksheet, sheet_raw)
            data_validators: List[DataValidation] = sheet_ws.data_validations.dataValidation

            dv_props_val = r.get("dv_props", [])
            dv_props = list(dv_props_val) if isinstance(dv_props_val, list) else []
            total_metric = len(data_validators) >= len(dv_props)
            for dat_vldt in data_validators:
                metric = False
                for prpt_item in dv_props:
                    prpt = cast(Rule, prpt_item)
                    metric = metric or all(
                        _match_value_to_rule(getattr(dat_vldt, attrbt, None), mr)  # type: ignore
                        for attrbt, mr in prpt.items()
                    )
                    if metric:
                        break
                total_metric = total_metric and metric
                if not total_metric:
                    break

            logger.debug("Assertion: %s.data_validation - %s", sheet_idx, total_metric)
            metric: bool = total_metric

        elif rule_type == "row_props":
            #  Check Row Properties
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            rows1: Dict[str, Union[float, bool, int]] = load_rows_or_cols(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke),
                obj="row",
                **r,  # type: ignore
            )
            rows2: Dict[str, Union[float, bool, int]] = load_rows_or_cols(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke),
                obj="row",
                **r,  # type: ignore
            )
            logger.debug("Rows1: %s", repr(rows1))
            logger.debug("Rows2: %s", repr(rows2))
            metric: bool = rows1 == rows2
            logger.debug(
                "Assertion: %s[rows] == %s[rows] - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "col_props":
            #  Check Row Properties
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            cols1: Dict[str, Union[float, bool, int]] = load_rows_or_cols(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke),
                obj="column",
                **r,  # type: ignore
            )
            cols2: Dict[str, Union[float, bool, int]] = load_rows_or_cols(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke),
                obj="column",
                **r,  # type: ignore
            )
            metric: bool = cols1 == cols2
            logger.debug(
                "Assertion: %s[cols] == %s[cols] - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "filter":
            #  Compare Filters
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            filters1: Dict[str, Union[str, List[str]]] = load_filters(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            filters2: Dict[str, Union[str, List[str]]] = load_filters(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            metric: bool = filters1 == filters2
            logger.debug(
                "Assertion: %s[filter] == %s[filter] - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "pivot_table":
            #  Compare Pivot Tables
            sheet_idx0 = cast(Union[int, str], r.get("sheet_idx0", 0))
            sheet_idx1 = cast(Union[int, str], r.get("sheet_idx1", 0))
            pivots1: Dict[str, Union[str, List[str]]] = load_pivot_tables(
                *parse_idx(sheet_idx0, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            pivots2: Dict[str, Union[str, List[str]]] = load_pivot_tables(
                *parse_idx(sheet_idx1, xlworkbookr, xlworkbooke),
                **r,  # type: ignore
            )
            metric: bool = pivots1 == pivots2
            logger.debug(
                "Assertion: %s[pivot]==%s[pivot] - %s",
                sheet_idx0,
                sheet_idx1,
                metric,
            )

        elif rule_type == "check_cell":
            #  Check Cell Properties
            try:
                sheet_idx = cast(Union[int, str], r.get("sheet_idx", 0))
                sheet_raw = _load_sheet(
                    *parse_idx(sheet_idx, xlworkbookr, xlworkbooke)  # type: ignore
                )
                if sheet_raw is None or not isinstance(sheet_raw, Worksheet):
                    logger.error(f"Failed to load sheet for sheet_idx: {sheet_idx}")
                    return 0.0
                sheet_ws = cast(Worksheet, sheet_raw)
                coordinate_val = r.get("coordinate", "A1")
                coordinate = str(coordinate_val)
                cell: Cell = sheet_ws[coordinate]
                metric: bool = True
                props_val = r.get("props", {})
                props = cast(Rule, props_val) if isinstance(props_val, dict) else {}
                for prpt, rule in props.items():
                    if prpt == "value":
                        try:
                            parsed_result = parse_idx(sheet_idx, result, expected)  # type: ignore
                            logger.debug(f"parse_idx result: {parsed_result}")
                            val = read_cell_value(
                                cast(str, parsed_result[0]), parsed_result[1], coordinate
                            )
                            logger.debug(f"Cell {coordinate} value: {val}")
                        except Exception as e:
                            logger.error(f"Failed to read cell value at {coordinate}: {e}")
                            val = None
                    else:
                        try:
                            val = _read_cell_style(prpt, cell)
                        except Exception as e:
                            logger.error(f"Failed to read cell style {prpt} at {coordinate}: {e}")
                            val = None

                    metric = metric and _match_value_to_rule(val, cast(Rule, rule))  # type: ignore
            except Exception as e:
                logger.error(f"Error in check_cell processing: {e}")
                return 0.0

            logger.debug(
                "Assertion: %s[%s] :%s - %s",
                sheet_idx,
                coordinate,
                repr(props),
                metric,
            )

        else:
            raise NotImplementedError("Unimplemented sheet check: {:}".format(r["type"]))

        passes = passes and metric
        if not passes:
            break

    return float(passes)


def _normalize_city_string(value: Union[str, int, float, None]) -> str:
    """Lowercase, strip punctuation, and remove accents for tolerant matching."""
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized.lower())
    return normalized.strip()


def compare_conference_city_in_order(
    actual_city_list_path: str, expected_city: Dict[str, List[Union[str, List[str]]]]
) -> float:
    """Compare conference city lists in order."""
    expected_city_list = expected_city["expected"]
    wb = openpyxl.load_workbook(actual_city_list_path)
    sheet = wb.active
    if sheet is None:
        logger.error("No active sheet found in workbook")
        return 0.0
    actual_city_list: List[Optional[str]] = []
    for row in sheet["C2:C22"]:
        for cell in row:
            actual_city_list.append(cell.value)

    try:
        for i, actual_city in enumerate(actual_city_list):
            actual_normalized = _normalize_city_string(actual_city)
            expected_entry = expected_city_list[i]

            if isinstance(expected_entry, str):
                expected_candidates = [expected_entry]
            elif isinstance(expected_entry, List):
                expected_candidates = expected_entry
            else:
                raise TypeError("Expected city should be a string or a list of strings")

            matched = False
            for candidate in expected_candidates:
                normalized_candidate = _normalize_city_string(candidate)
                if normalized_candidate and normalized_candidate in actual_normalized:
                    matched = True
                    break

            if not matched:
                logger.debug(f"Expected city {expected_entry}; Actual city {actual_city}")
                print(f"Expected city {expected_entry}; Actual city {actual_city}")
                return 0.0

    except Exception as exc:
        logger.error(f"Error comparing conference cities: {exc}")
        return 0.0

    return 1.0
