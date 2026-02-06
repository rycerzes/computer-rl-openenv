import builtins
import difflib
import functools
import itertools
import logging
import operator
import os
import re
import zipfile
from typing import (
    Callable,
    Dict,
    Iterable,
    List,
    Match,
    Optional,
    Pattern,
    Set,
    Tuple,
    TypeVar,
    Union,
    cast,
)
from urllib.parse import ParseResult, urlparse, urlunparse

import formulas
import lxml.cssselect
import lxml.etree
import tldextract
import xmltodict
from lxml.etree import _Element
from openpyxl import Workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.chart._chart import ChartBase
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.pivot.cache import CacheSource as PivotCacheSource
from openpyxl.pivot.table import TableDefinition as PivotTableDefinition
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.worksheet.filters import AutoFilter, SortState
from openpyxl.worksheet.worksheet import Worksheet

V = TypeVar("V")

logger = logging.getLogger("computer_rl_env.server.evaluators.metrics.utils")

_xlsx_namespaces = [
    ("oo", "http://schemas.openxmlformats.org/spreadsheetml/2006/main"),
    ("x14", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"),
    ("xm", "http://schemas.microsoft.com/office/excel/2006/main"),
]
_xlsx_ns_mapping = dict(_xlsx_namespaces)
_xlsx_ns_imapping: Dict[str, Optional[str]] = dict(
    map(lambda itm: (itm[1], itm[0]), _xlsx_namespaces)
)
_xlsx_ns_imapping["http://schemas.openxmlformats.org/spreadsheetml/2006/main"] = None
_sheet_name_selector = lxml.cssselect.CSSSelector("oo|sheets>oo|sheet", namespaces=_xlsx_ns_mapping)
_sparklines_selector = lxml.cssselect.CSSSelector("x14|sparkline", namespaces=_xlsx_ns_mapping)


def load_sparklines(xlsx_file: str, sheet_name: str) -> Dict[str, str]:
    """
    Args:
        xlsx_file: path to xlsx
        sheet_name: sheet name

    Returns:
        sparkline definitions in form of {"F3": "Sheet1!C3:E3"}
    """
    try:
        with zipfile.ZipFile(xlsx_file, "r") as z_f:
            with z_f.open("xl/workbook.xml") as f:
                workbook_database: _Element = lxml.etree.fromstring(f.read())
                sheets: List[_Element] = _sheet_name_selector(workbook_database)
                sheet_names: Dict[str, str] = {sh.get("name"): sh.get("sheetId") for sh in sheets}
            with z_f.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as f:
                sheet: _Element = lxml.etree.fromstring(f.read())
                sparklines: List[_Element] = _sparklines_selector(sheet)
    except zipfile.BadZipFile:
        return {}

    sparklines_dict: Dict[str, str] = {}
    for sp_l in sparklines:
        sparkline_xml: str = lxml.etree.tostring(sp_l, encoding="unicode")
        sparkline: Dict[str, Dict[str, str]] = xmltodict.parse(
            sparkline_xml,
            process_namespaces=True,
            namespaces=_xlsx_ns_imapping,  # type: ignore[arg-type]
        )
        sparklines_dict[sparkline["x14:sparkline"]["xm:sqref"]] = sparkline["x14:sparkline"]["xm:f"]
    return sparklines_dict


def load_charts(
    xlsx_file: Workbook, sheet_name: str, **options
) -> Dict[str, Dict[str, Union[str, List[Union[str, int, float]], int, float, None]]]:
    """
    Args:
        xlsx_file: concerned excel book
        sheet_name: sheet name
        options: dict like {"chart_props": list of str} giving the concerned chart properties

    Returns:
        information of charts, dict like {<str representing data source>: {<str as property>: value}}
    """
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    charts: List[ChartBase] = worksheet._charts  # type: ignore[attr-defined]

    chart_set: Dict[str, Dict[str, Union[str, List[Union[str, int, float]], int, float, None]]] = {}
    chart_props: Set[str] = set(options["chart_props"]) if "chart_props" in options else set()
    for ch in charts:
        series: List[str] = []
        for ser in ch.series:
            if hasattr(ser.val, "numRef") and hasattr(ser.val.numRef, "f"):
                value_str: str = ser.val.numRef.f
            elif hasattr(ser.val, "strRef") and hasattr(ser.val.strRef, "f"):
                value_str: str = ser.val.strRef.f
            else:
                value_str: str = ""
            if hasattr(ser.cat, "numRef") and hasattr(ser.cat.numRef, "f"):
                categ_str: str = ser.cat.numRef.f
            elif hasattr(ser.cat, "strRef") and hasattr(ser.cat.strRef, "f"):
                categ_str: str = ser.cat.strRef.f
            else:
                categ_str: str = ""
            series.append("{:},{:}".format(value_str, categ_str))
        series_str: str = ";".join(series)

        info: Dict[str, Union[str, List[Union[str, int, float]], int, float, None]] = {}

        if "title" in chart_props:
            try:
                info["title"] = ch.title.tx.rich.p[0].r[0].t  # type: ignore[union-attr, index]
            except Exception:
                info["title"] = None
        if "legend" in chart_props:
            info["legend"] = ch.legend.position if ch.legend is not None else None
        if "anchor" in chart_props:
            info["anchor"] = [  # type: ignore[union-attr]
                ch.anchor.editAs,  # type: ignore[union-attr]
                ch.anchor._from.col,  # type: ignore[union-attr]
                ch.anchor._from.row,  # type: ignore[union-attr]
                ch.anchor.to.col,  # type: ignore[union-attr]
                ch.anchor.to.row,  # type: ignore[union-attr]
            ]
        if "width" in chart_props:
            info["width"] = ch.width
        if "height" in chart_props:
            info["height"] = ch.height
        if "type" in chart_props:
            info["type"] = ch.tagname
        if "direction" in chart_props:
            info["direction"] = ch.barDir  # type: ignore[attr-defined]

        if "xtitle" in chart_props:
            try:
                info["xtitle"] = ch.x_axis.title.tx.rich.p[0].r[0].t  # type: ignore[attr-defined, union-attr, index]
            except Exception:
                info["xtitle"] = None
        if "ytitle" in chart_props:
            try:
                info["ytitle"] = ch.y_axis.title.tx.rich.p[0].r[0].t  # type: ignore[attr-defined, union-attr, index]
            except Exception:
                info["ytitle"] = None
        if "ztitle" in chart_props:
            try:
                info["ztitle"] = ch.z_axis.title.tx.rich.p[0].r[0].t  # type: ignore[attr-defined, union-attr, index]
            except Exception:
                info["ztitle"] = None
        chart_set[series_str] = info
    logger.debug(".[%s].charts: %s", sheet_name, repr(chart_set))
    return chart_set


def load_pivot_tables(
    xlsx_file: Workbook, sheet_name: str, **options
) -> Dict[
    str,
    Dict[str, Union[str, bool, List[Union[int, str, List[Tuple[Optional[bool], int]]]], Set[int]]],
]:
    """
    Args:
        xlsx_file: concerned excel book
        sheet_name: sheet name
        options: dict like {"pivot_props": list of str} giving the concerned pivot properties

    Returns:
        information of pivot tables
    """
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    pivots: List[PivotTableDefinition] = worksheet._pivots  # type: ignore[attr-defined]

    pivot_set: Dict[
        str,
        Dict[
            str, Union[str, bool, List[Union[int, str, List[Tuple[Optional[bool], int]]]], Set[int]]
        ],
    ] = {}
    pivot_props: Set[str] = set(options.get("pivot_props", []))
    for pvt in pivots:
        raw_selection: List[List[Tuple[Optional[bool], int]]] = [
            [(itm.h, itm.x) for itm in f.items if itm.x is not None] for f in pvt.pivotFields
        ]
        raw__selection: List[List[Tuple[Optional[bool], int]]] = list(
            itertools.dropwhile(lambda r: len(r) == 0, raw_selection)
        )
        left_bias = len(raw_selection) - len(raw__selection)
        selection: List[List[Tuple[Optional[bool], int]]] = list(
            (itertools.dropwhile(lambda r: len(r) == 0, reversed(raw__selection)))
        )[::-1]
        right_bias = len(raw__selection) - len(selection)
        cache_source: PivotCacheSource = pvt.cache.cacheSource
        cell_range1: str
        cell_range2: str
        cell_range1, cell_range2 = cache_source.worksheetSource.ref.split(":")  # type: ignore[union-attr]
        cell_range1_tuple: Tuple[int, int] = coordinate_to_tuple(cell_range1)
        cell_range1_tuple = (cell_range1_tuple[0], cell_range1_tuple[1] + left_bias)
        cell_range2_tuple: Tuple[int, int] = coordinate_to_tuple(cell_range2)
        cell_range2_tuple = (cell_range2_tuple[0], cell_range2_tuple[1] - right_bias)
        source: str = "{:};{:}:{:};{:}".format(
            cache_source.type,
            cell_range1_tuple,
            cell_range2_tuple,
            cache_source.worksheetSource.sheet,  # type: ignore[union-attr]
        )

        info: Dict[
            str, Union[str, bool, List[Union[int, str, List[Tuple[Optional[bool], int]]]], Set[int]]
        ] = {}
        if "name" in pivot_props:
            info["name"] = pvt.name

        if "show_total" in pivot_props:
            info["show_total"] = pvt.visualTotals
        if "show_empty_row" in pivot_props:
            info["show_empty_row"] = pvt.showEmptyRow
        if "show_empty_col" in pivot_props:
            info["show_empty_col"] = pvt.showEmptyCol
        if "show_headers" in pivot_props:
            info["show_headers"] = pvt.showHeaders

        if "location" in pivot_props:
            info["location"] = str(pvt.location)  # type: ignore[arg-type]
        if "filter" in pivot_props or "selection" in pivot_props:
            info["selection"] = (  # type: ignore[assignment]
                selection if "ordered" in pivot_props else [set(r) for r in selection]
            )
        if "filter" in pivot_props:
            info["filter_fields"] = set(f.fld for f in pvt.pageFields)
        if "col_fields" in pivot_props:
            info["col_fields"] = [f.x - left_bias for f in pvt.colFields]
        if "row_fields" in pivot_props:
            info["row_fields"] = [f.x - left_bias for f in pvt.rowFields]
        if "data_fields" in pivot_props:
            info["data_fields"] = [
                "{:d};{:};{:};{:}".format(
                    f.fld - left_bias,
                    f.name if "data_fields_name" in pivot_props else "",
                    f.subtotal,
                    f.showDataAs,
                )
                for f in pvt.dataFields
            ]

        pivot_set[source] = info
    logger.debug(".[%s].pivots: %s", sheet_name, repr(pivot_set))
    return pivot_set


_shared_str_selector = lxml.cssselect.CSSSelector("oo|sst>oo|si", namespaces=_xlsx_ns_mapping)
_shared_str_value_selector = lxml.cssselect.CSSSelector("oo|t", namespaces=_xlsx_ns_mapping)


def read_cell_value(xlsx_file: str, sheet_name: str, coordinate: str) -> Union[str, float, None]:
    """Read cell value from xlsx file."""
    logger.debug(
        f"Reading cell value from {xlsx_file}, sheet: {sheet_name}, coordinate: {coordinate}"
    )

    if not os.path.exists(xlsx_file):
        logger.error(f"Excel file not found: {xlsx_file}")
        return None

    try:
        with zipfile.ZipFile(xlsx_file, "r") as z_f:
            try:
                with z_f.open("xl/sharedStrings.xml") as f:
                    shared_str_xml: _Element = lxml.etree.fromstring(f.read())
                    str_elements: List[_Element] = _shared_str_selector(shared_str_xml)
                    shared_strs: List[str] = [
                        "".join(t.text for t in _shared_str_value_selector(elm))
                        for elm in str_elements
                    ]
            except Exception:
                logger.debug("Read shared strings error: %s", xlsx_file)
                shared_strs: List[str] = []

            with z_f.open("xl/workbook.xml") as f:
                workbook_database: _Element = lxml.etree.fromstring(f.read())
                sheets: List[_Element] = _sheet_name_selector(workbook_database)
                sheet_names: Dict[str, str] = {sh.get("name"): sh.get("sheetId") for sh in sheets}

            with z_f.open("xl/worksheets/sheet{:}.xml".format(sheet_names[sheet_name])) as f:
                sheet: _Element = lxml.etree.fromstring(f.read())
                cells: List[_Element] = lxml.cssselect.CSSSelector(
                    'oo|row>oo|c[r="{:}"]'.format(coordinate), namespaces=_xlsx_ns_mapping
                )(sheet)
                if len(cells) == 0:
                    logger.debug(f"Cell {coordinate} not found in sheet {sheet_name}")
                    return None
                cell: _Element = cells[0]
    except zipfile.BadZipFile as e:
        logger.error(f"Bad zip file {xlsx_file}: {e}")
        return None
    except KeyError as e:
        logger.error(f"Sheet {sheet_name} not found in {xlsx_file}: {e}")
        return None
    except Exception as e:
        logger.error(f"Error reading {xlsx_file}: {e}")
        return None

    cell_dict: Dict = xmltodict.parse(  # type: ignore[type-arg]
        lxml.etree.tostring(cell, encoding="unicode"),
        process_namespaces=True,
        namespaces=_xlsx_ns_imapping,  # type: ignore[arg-type]
    )
    logger.debug("%s.shared_strings: %s", xlsx_file, repr(shared_strs))
    logger.debug("%s.%s[%s]: %s", xlsx_file, sheet_name, coordinate, repr(cell_dict))
    try:
        if "@t" not in cell_dict["c"] or cell_dict["c"]["@t"] == "n":  # type: ignore[literal-required]
            return float(cell_dict["c"]["v"])  # type: ignore[literal-required]
        if cell_dict["c"]["@t"] == "s":  # type: ignore[literal-required]
            return shared_strs[int(cell_dict["c"]["v"])]  # type: ignore[literal-required]
        if cell_dict["c"]["@t"] == "str":  # type: ignore[literal-required]
            return cell_dict["c"]["v"]  # type: ignore[literal-required]
        if cell_dict["c"]["@t"] == "inlineStr":  # type: ignore[literal-required]
            return cell_dict["c"]["is"]["t"]  # type: ignore[literal-required]
        if cell_dict["c"]["@t"] == "e":  # type: ignore[literal-required]
            return cell_dict["c"]["v"]  # type: ignore[literal-required]
    except (KeyError, ValueError):
        return None
    return None


def _read_cell_style(
    style_name: str, cell: Union[Cell, MergedCell], diff_style: Optional[DifferentialStyle] = None
) -> Union[str, float, bool, None]:
    """Read cell style property."""
    if style_name == "number_format":
        return (
            (cell.number_format if diff_style is None else diff_style.numFmt.formatCode)  # type: ignore[union-attr]
            if cell.value is not None and cell.data_type == "n"
            else None
        )
    elif style_name == "font_name":
        return (diff_style or cell).font.name if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_family":
        return (diff_style or cell).font.family if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_color":
        return (diff_style or cell).font.color.rgb if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_bold":
        return (diff_style or cell).font.bold if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_italic":
        return (diff_style or cell).font.italic if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_underline":
        return (diff_style or cell).font.underline if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "font_size":
        return (diff_style or cell).font.size if cell.value is not None else None  # type: ignore[union-attr]
    elif style_name == "fill_type":
        try:
            return (diff_style or cell).fill.tagname  # type: ignore[union-attr]
        except Exception:
            return None
    elif style_name == "bgcolor" or style_name == "fgcolor":
        try:
            if diff_style is not None:
                return diff_style.fill.bgColor.rgb  # type: ignore[union-attr]
            else:
                return cell.fill.fgColor.rgb  # type: ignore[union-attr]
        except Exception:
            return None
    elif style_name == "hyperlink":
        link = cell.hyperlink
        return (
            str(link.target)
            if link and hasattr(link, "target")
            else ("" if cell.value is not None else None)
        )
    elif style_name == "merge":
        return isinstance(cell, MergedCell)
    else:
        raise NotImplementedError("Unsupported Style: {:}".format(style_name))


def _process_xlsx_cf_operator(
    operator: str, value: Union[str, float, int], ref: List[Union[str, float, int]]
) -> bool:
    """Process xlsx conditional formatting operator."""
    try:
        if operator == "lessThanOrEqual":
            result: bool = value <= ref[0]  # type: ignore[operator]
        elif operator == "lessThan":
            result: bool = value < ref[0]  # type: ignore[operator]
        elif operator == "equal":
            result: bool = value == ref[0]
        elif operator == "greaterThanOrEqual":
            result: bool = value >= ref[0]  # type: ignore[operator]
        elif operator == "notEqual":
            result: bool = value != ref[0]
        elif operator == "greaterThan":
            result: bool = value > ref[0]  # type: ignore[operator]
        elif operator == "between":
            small_one: Union[str, float, int] = min(ref)
            large_one: Union[str, float, int] = max(ref)
            result: bool = value >= small_one and value <= large_one  # type: ignore[operator]
        elif operator == "notBetween":
            small_one: Union[str, float, int] = min(ref)
            large_one: Union[str, float, int] = max(ref)
            result: bool = value < small_one or value > large_one  # type: ignore[operator]
        else:
            logger.exception("Not Implemented CondFormat Operator: {:}".format(operator))
            result = False
        return result
    except TypeError:
        logger.exception("Unmatched type of %s and %s. Auto to False", repr(value), repr(ref))
        return False
    except IndexError:
        logger.exception("ref array doesn't have enough elements. Auto to False: %s", repr(ref))
        return False


_absolute_range_pattern: Pattern[str] = re.compile(
    r"""\$(?P<col1>[A-Z]{1,3})\$(?P<row1>\d+)  # coord1
        (?::
          \$(?P<col2>[A-Z]{1,3})\$(?P<row2>\d+)  # coord2
        )?
    """,
    re.X,
)


def load_xlsx_styles(
    xlsx_file: Workbook, sheet_name: str, book_name: str, **options
) -> Dict[str, List[Union[str, float, bool, None]]]:
    """
    Args:
        xlsx_file: concerned excel book
        sheet_name: sheet name
        book_name: book name
        options: dict like {"props": list of str} giving the concerned styles

    Returns:
        dict like {<str as cell coordinates>: list of property values}
    """
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    style_dict: Dict[str, List[Union[str, float, bool, None]]] = {}
    concerned_styles: List[str] = options.get("props", [])

    # Handles Cell Styles
    for col in worksheet.iter_cols():
        for c in col:
            style_list: List[Union[str, float, bool, None]] = []
            for st in concerned_styles:
                style_list.append(_read_cell_style(st, c))
            style_dict[c.coordinate] = style_list

    # Handles Conditional Formatting
    conditional_formattings: ConditionalFormattingList = worksheet.conditional_formatting
    formula_parser = formulas.Parser()
    for fmt in conditional_formattings:
        for r in fmt.rules:
            active_cells: List[Cell] = []

            # Process CF Formulae
            formulae: List[Callable] = []
            argument_lists: List[List[Union[str, float, List[Union[str, float, None]]]]] = []
            has_error = False
            for fml in r.formula:
                try:
                    formula_func: Callable = formula_parser.ast("=" + fml)[1].compile()  # type: ignore[misc]
                    logger.debug("CondFormat rule formula: %s", fml)
                except Exception:
                    logger.exception("Formula parsing error: %s. Skipping.", repr(fml))
                    has_error = True
                    break

                arguments: List[Union[str, float, List[Union[str, float, None]]]] = []
                absolute_range_match: List[Tuple[str, str, str, str]] = (
                    _absolute_range_pattern.findall(fml)
                )
                for m in absolute_range_match:
                    logger.debug("Absolute ranges: %s", repr(m))
                    if m[2] == "" and m[3] == "":
                        cell_val = read_cell_value(
                            book_name, sheet_name, coordinate="{:}{:}".format(m[0], m[1])
                        )
                        if cell_val is not None:
                            arguments.append(cell_val)
                    else:
                        arguments.append(
                            [
                                read_cell_value(
                                    book_name,
                                    sheet_name,
                                    coordinate="{:}{:}".format(get_column_letter(c[1]), c[0]),
                                )
                                for c in CellRange(
                                    "{:}{:}:{:}{:}".format(m[0], m[1], m[2], m[3])
                                ).cells
                            ]
                        )
                logger.debug("Absolute range arguments: %s", repr(arguments))

                formulae.append(formula_func)
                argument_lists.append(arguments)

            if has_error:
                continue

            # Process Condition According to Type
            if r.type in {
                "expression",
                "containsText",
                "notContainsText",
                "endsWith",
                "beginsWith",
                "containsErrors",
                "notContainsErrors",
            }:
                condition: Callable = formulae[0]  # type: ignore[misc]
                arguments_for_cond: List[Union[str, float, List[Union[str, float, None]]]] = (
                    argument_lists[0]
                )

                def is_active(v: Union[str, float]) -> bool:
                    return condition(v, *arguments_for_cond)  # type: ignore[misc]

            elif r.type == "cellIs":
                operator_str: str = r.operator
                try:
                    references: List[Union[str, float, int]] = [fml() for fml in formulae]  # type: ignore[misc]
                except Exception:
                    logger.exception(
                        "Error occurs while calculating reference values for cellIs condition formatting."
                    )
                    continue

                def is_active(v: Union[str, float]) -> bool:
                    return _process_xlsx_cf_operator(operator_str, v, references)
            else:
                logger.exception("Not Implemented Condition Type: {:}".format(r.type))
                continue

            # Test Each Cell
            nb_contiguous_nothings = 0
            for rge in fmt.cells:
                for c in rge.cells:
                    cell_obj = worksheet.cell(row=c[0], column=c[1])
                    cell_value = read_cell_value(
                        book_name,
                        sheet_name,
                        coordinate="{:}{:d}".format(get_column_letter(c[1]), c[0]),
                    )
                    if cell_value is None:
                        nb_contiguous_nothings += 1
                        if nb_contiguous_nothings > 50:
                            break
                        continue
                    else:
                        try:
                            satisfies_condition: bool = is_active(cell_value)
                        except Exception:
                            logger.exception(
                                "Error in formula calculation with cell value %s", repr(cell_value)
                            )
                            satisfies_condition = False
                        if satisfies_condition:
                            logger.debug(
                                "Active Cell %s(%s) for %s",
                                repr(cell_obj),
                                repr(cell_value),
                                r.formula[0],
                            )
                            active_cells.append(cell_obj)  # type: ignore[arg-type]

            for c in active_cells:
                style_dict[c.coordinate] = [
                    _read_cell_style(st, c, r.dxf) for st in concerned_styles
                ]

    logger.debug(".[%s].styles: %s", sheet_name, repr(style_dict))
    return style_dict


def load_rows_or_cols(
    xlsx_file: Workbook, sheet_name: str, **options
) -> Dict[int, Dict[str, Union[float, bool, int]]]:
    """
    Args:
        xlsx_file: concerned excel book
        sheet_name: sheet name
        options: dict like {"obj": "row" | "column", "props": list of str} giving the concerned row/column properties

    Returns:
        row/column information
    """
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}
    objs: DimensionHolder = getattr(worksheet, "{:}_dimensions".format(options["obj"]))

    obj_set: Dict[int, Dict[str, Union[float, bool, int]]] = {}
    obj_props: Set[str] = set(options.get("props", []))
    for obj_no, obj_dms in objs.items():
        info_dict: Dict[str, Union[float, bool, int]] = {}
        for prop in obj_props:
            info_dict[prop] = getattr(obj_dms, prop)
        obj_set[obj_no] = info_dict
    return obj_set


def load_filters(
    xlsx_file: Workbook, sheet_name: str, **options
) -> Dict[
    str,
    Union[
        str,
        List[Dict[str, Union[int, bool, Set[Union[str, Tuple[str, str]]]]]],
        Dict[str, Union[bool, str, List[Dict[str, Union[bool, str, int, None]]], None]],
    ],
]:
    """Load filter information from xlsx file."""
    try:
        worksheet: Worksheet = xlsx_file[sheet_name]
    except KeyError:
        return {}

    filters: AutoFilter = worksheet.auto_filter
    filter_dict: Dict[
        str,
        Union[
            str,
            List[Dict[str, Union[int, bool, Set[Union[str, Tuple[str, str]]]]]],
            Dict[str, Union[bool, str, List[Dict[str, Union[bool, str, int, None]]], None]],
        ],
    ] = {}
    filter_dict["ref"] = filters.ref

    # filterColumn
    filter_column_set: List[Dict[str, Union[int, bool, Set[Union[str, Tuple[str, str]]]]]] = []
    for flt_clm in filters.filterColumn:
        filter_column: Dict[str, Union[int, bool, Set[Union[str, Tuple[str, str]]]]] = {}
        filter_column["col_id"] = flt_clm.colId
        filter_column["hidden_button"] = flt_clm.hiddenButton
        filter_column["show_button"] = flt_clm.showButton
        if flt_clm.filters is not None:
            filter_column["filters_blank"] = flt_clm.filters.blank
            filter_column["filters"] = set(flt_clm.filters.filter)
        if flt_clm.customFilters is not None:
            filter_column["custom_filters_op"] = flt_clm.customFilters._and
            filter_column["custom_filters"] = set(
                (flt.operator, flt.val) for flt in flt_clm.customFilters.customFilter
            )
        filter_column_set.append(filter_column)
    filter_column_set = list(sorted(filter_column_set, key=(lambda d: d["col_id"])))
    filter_dict["filter_column"] = filter_column_set

    # sortState
    sort_state: Optional[SortState] = filters.sortState
    if sort_state is not None:
        sort_state_dict: Dict[
            str, Union[bool, str, List[Dict[str, Union[bool, str, int, None]]], None]
        ] = {}
        sort_state_dict["sort"] = sort_state.columnSort
        sort_state_dict["case"] = sort_state.caseSensitive
        sort_state_dict["method"] = sort_state.sortMethod  # type: ignore[assignment]
        sort_state_dict["ref"] = sort_state.ref
        sort_state_dict["condition"] = list(
            {
                "descending": cdt.descending,
                "key": cdt.sortBy,
                "ref": cdt.ref,
                "custom_list": cdt.customList,
                "dxf_id": cdt.dxfId,
                "icon": cdt.iconSet,
                "iconid": cdt.iconId,
            }
            for cdt in sort_state.sortCondition
        )
        filter_dict["sort_state"] = sort_state_dict

    return filter_dict


def _match_record(
    pattern: Dict[str, Union[str, int, float, bool]], item: Dict[str, Union[str, int, float, bool]]
) -> bool:
    """Match a pattern dict against an item dict."""
    return all(k in item and item[k] == val for k, val in pattern.items())


def _multicellrange_containsby(
    subset_candidate: MultiCellRange, superset_candidate: MultiCellRange
) -> bool:
    """Check if subset is contained by superset."""
    return all(r in superset_candidate for r in subset_candidate)


def _match_value_to_rule(value: V, rule: Dict[str, Union[str, V]]) -> bool:
    """
    Args:
        value: value to match
        rule: rule dict like {"method": str, "ref": V as ref value}

    Returns:
        bool
    """
    method = cast(str, rule["method"])
    if method.startswith("re"):  # re.FLAGs
        flags: List[str] = method.split(".")[1:]
        flags_iter: Iterable[re.RegexFlag] = (getattr(re, fl) for fl in flags)
        flag: re.RegexFlag = functools.reduce(operator.or_, flags_iter, re.RegexFlag(0))
        logger.debug("REFLAG: %s", repr(flag))

        match_: Optional[Match[str]] = re.search(cast(str, rule["ref"]), str(value), flag)
        return match_ is not None
    if method in {"eq", "ne", "le", "lt", "ge", "gt"}:
        return getattr(operator, method)(value, rule["ref"])  # type: ignore[arg-type, operator]
    if method.startswith("approx"):  # approx:THRESHOLD
        threshold: float = float(method.split(":")[1])
        logger.debug("Approx: TH%f, REF%f, VAL%s", threshold, rule["ref"], repr(value))
        try:
            value_float: float = float(value)  # type: ignore[arg-type]
        except (ValueError, TypeError):
            return False
        else:
            return abs(value_float - cast(float, rule["ref"])) <= threshold
    if method == "spreadsheet_range":
        ref_list = cast(List, rule["ref"])
        subset_limit = MultiCellRange(ref_list[0])
        superset_limit = MultiCellRange(ref_list[1])
        return _multicellrange_containsby(
            subset_limit, cast(MultiCellRange, value)
        ) and _multicellrange_containsby(cast(MultiCellRange, value), superset_limit)
    if method.startswith("range."):  # e.g., range.te [0, 2] -> 0 < x <= 2
        left_et = method[6]
        right_et = method[7]
        ref_list = cast(List, rule["ref"])
        return getattr(operator, "l" + left_et)(ref_list[0], value) and getattr(  # type: ignore[arg-type, operator]
            operator, "l" + right_et
        )(value, ref_list[1])  # type: ignore[arg-type, operator]
    if method in {"str_list_eq", "str_set_eq"}:
        container_type_str: str = method[4:-3]
        container_type = getattr(builtins, container_type_str)

        value_container = container_type(str(value).strip("\"'").split(","))
        ref_container = container_type(rule["ref"])
        return value_container == ref_container
    raise NotImplementedError()


def are_lists_equal(
    list1: List, list2: List, comparison_func: Callable[[object, object], bool]
) -> bool:
    """Check if two lists are equal using a comparison function."""
    if len(list1) != len(list2):
        return False

    for item1 in list1:
        if not any(comparison_func(item1, item2) for item2 in list2):
            return False

    return True


def compare_urls(url1: Optional[str], url2: Optional[str], full: bool = True) -> bool:
    """Compare two URLs for similarity."""
    if url1 is None or url2 is None:
        return url1 == url2

    logger.info(f"compare_urls. url1: {url1}; url2: {url2}")

    def parse_with_default_scheme(url: str) -> ParseResult:
        """Ensure the URL has a scheme. If not, prepend 'http://'."""
        if not re.match(r"^[a-zA-Z][a-zA-Z0-9+\-.]*://", url):
            url = f"http://{url}"
        return urlparse(url)

    def normalize_url(url: str) -> str:
        parsed_url = parse_with_default_scheme(url)
        scheme = parsed_url.scheme.lower()

        extracted = tldextract.extract(parsed_url.netloc.lower())

        subdomain = extracted.subdomain
        if subdomain == "www":
            subdomain = ""

        if subdomain:
            normalized_netloc = f"{subdomain}.{extracted.domain}"
        else:
            normalized_netloc = extracted.domain

        normalized_path = parsed_url.path if parsed_url.path != "/" else ""

        normalized_parsed_url = ParseResult(
            scheme=scheme.lower(),
            netloc=normalized_netloc,
            path=normalized_path,
            params=parsed_url.params if full else "",
            query=parsed_url.query if full else "",
            fragment=parsed_url.fragment if full else "",
        )
        return urlunparse(normalized_parsed_url)

    logger.info(f"After normalization. url1: {normalize_url(url1)}; url2: {normalize_url(url2)}")
    norm_url1 = normalize_url(url1)
    norm_url2 = normalize_url(url2)

    return norm_url1 == norm_url2


def exact_match(actual: Union[str, int, float], expected: Union[str, int, float]) -> bool:
    """Check exact match between actual and expected values."""
    return str(actual).strip() == str(expected).strip()


def fuzzy_match(actual: Union[str, int, float], expected: Union[str, int, float]) -> float:
    """Calculate fuzzy match ratio between actual and expected values."""
    return difflib.SequenceMatcher(None, str(actual), str(expected)).ratio()
